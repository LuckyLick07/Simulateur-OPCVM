#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Intègre les historiques de VL dans le simulateur.

Lit chaque fichier de donnees-vl/ (CSV, ou Excel si openpyxl est installé),
en extrait les couples (date, VL), puis engendre index.html à la racine du
dépôt à partir de outil/gabarit.html (remplacement du repère __FONDSDATA__).

Usage :  python3 outil/integrer_vl.py        (depuis la racine du dépôt)
         python3 integrer_vl.py              (depuis outil/ — équivalent)
"""

import json
import math
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path

RACINE = Path(__file__).resolve().parent.parent
DOSSIER_VL = RACINE / "donnees-vl"
GABARIT = RACINE / "outil" / "gabarit.html"
SORTIE = RACINE / "index.html"

# ————— Réglages par fonds (facultatifs) —————
# Clé : fragment (insensible à la casse) du nom de fichier.
# ordre : position dans la liste déroulante ; entree/sortie : droits en % ;
# politique : orientation de gestion affichée dans la fiche.
# rendement_du / rendement_au : fenêtre de calcul du rendement appliqué
# (borne = dernière cotation à la date ou avant) — l'historique complet reste
# affiché ; rendement_libelle : explication de la fenêtre dans la fiche.
REGLAGES = {
    "confort": {"ordre": 1, "entree": 0.75, "sortie": 0.0, "politique": "obligataire"},
    "capital": {"ordre": 2, "entree": 1.5, "sortie": 1.5, "politique": "diversifié"},
    "brvm": {"ordre": 9, "entree": 0.0, "sortie": 0.0,
             "type": "indice", "nom": "BRVM Composite (indice)",
             "rendement_du": "2020-12-31", "rendement_au": "2025-12-31",
             "rendement_libelle": "les cinq dernières années civiles (2021-2025)"},
}

# ————— Fonds récents sans historique de VL : rendement cible annoncé —————
# Tant qu'un fonds n'a pas d'historique, il est simulé au taux que la société
# de gestion s'est fixé comme objectif (étiqueté « cible, non garanti » partout
# sur le site, et exclu du parcours guidé). Dès que ses VL existent : déposer
# le CSV dans donnees-vl/ et retirer l'entrée ci-dessous — le fonds bascule
# alors sur le rendement observé. taux en % par an ; entree/sortie en %.
FONDS_CIBLES = [
    {"nom": "FCP AGA Elite", "politique": "actions",
     "taux": 15.0, "annee": 2026, "ordre": 3, "entree": 2.0, "sortie": 2.0},
    {"nom": "FCP AGA Obligations", "politique": "obligataire (« investment grade »)",
     "taux": 6.0, "annee": 2026, "ordre": 4, "entree": 1.0, "sortie": 1.0},
    {"nom": "FCP AGA TrésoActiv", "politique": "monétaire",
     "taux": 5.5, "annee": 2026, "ordre": 5, "entree": 0.5, "sortie": 0.5},
]

# ————— Envoi du résumé par e-mail (EmailJS) —————
# Laisser vide tant que le compte n'est pas configuré : le bouton bascule
# alors sur un brouillon dans la messagerie du visiteur.
EMAILJS = {
    "service": "service_bk2nicj",
    "template": "template_4sxn887",
    "publicKey": "DEtodI8QynsdM3YHw",
}


def lire_texte(chemin: Path) -> str:
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return chemin.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    raise SystemExit(f"Encodage illisible : {chemin.name}")


def parser_date(brut: str):
    brut = brut.strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", brut)  # JJ/MM/AAAA
    if m:
        j, mo, a = map(int, m.groups())
        return date(a, mo, j)
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", brut)  # AAAA-MM-JJ
    if m:
        a, mo, j = map(int, m.groups())
        return date(a, mo, j)
    return None


def parser_vl(brut: str):
    # « 1 602,82 » → 1602.82 (espaces fines, insécables et normales acceptées)
    brut = brut.strip().replace(" ", "").replace(" ", "").replace(" ", "")
    brut = brut.replace(",", ".")
    if not brut:
        return None
    try:
        v = float(brut)
    except ValueError:
        return None
    return v if v > 0 else None


def lire_csv(chemin: Path):
    texte = lire_texte(chemin)
    sep = ";" if ";" in texte.splitlines()[0] else ","
    couples = []
    for ligne in texte.splitlines():
        cols = [c.strip() for c in ligne.split(sep)]
        if len(cols) < 2:
            continue
        d = parser_date(cols[0])
        v = parser_vl(cols[1])
        if d is not None and v is not None:
            couples.append((d, v))
    return couples


def lire_xlsx(chemin: Path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            f"{chemin.name} : lire un fichier Excel demande openpyxl "
            "(pip install openpyxl), ou convertir le fichier en CSV."
        )
    feuille = load_workbook(chemin, data_only=True).worksheets[0]
    couples = []
    for rangee in feuille.iter_rows(values_only=True):
        if not rangee or rangee[0] is None or len(rangee) < 2:
            continue
        brut_d, brut_v = rangee[0], rangee[1]
        d = brut_d.date() if hasattr(brut_d, "date") else parser_date(str(brut_d))
        v = brut_v if isinstance(brut_v, (int, float)) else parser_vl(str(brut_v))
        if d is not None and v is not None and v > 0:
            couples.append((d, float(v)))
    return couples


def nom_du_fonds(chemin: Path) -> str:
    nom = chemin.stem
    nom = re.sub(r"^\s*historique\s+(VL\s+)?", "", nom, flags=re.IGNORECASE)
    mots = []
    for mot in nom.split():
        mots.append(mot if len(mot) <= 3 else mot.capitalize())
    return " ".join(mots)


def reglages_pour(chemin: Path) -> dict:
    nom_plat = unicodedata.normalize("NFKD", chemin.stem).encode("ascii", "ignore").decode().lower()
    for fragment, reglage in REGLAGES.items():
        if fragment in nom_plat:
            return reglage
    return {}


def main():
    fichiers = sorted(
        p for p in DOSSIER_VL.iterdir()
        if p.suffix.lower() in (".csv", ".xlsx") and not p.name.startswith(".")
    )
    if not fichiers:
        raise SystemExit(f"Aucun fichier de VL dans {DOSSIER_VL}")

    fonds = []
    for chemin in fichiers:
        couples = lire_xlsx(chemin) if chemin.suffix.lower() == ".xlsx" else lire_csv(chemin)
        # tri chronologique + dédoublonnage par date (dernière valeur retenue)
        par_date = {}
        for d, v in couples:
            par_date[d] = v
        couples = sorted(par_date.items())
        if len(couples) < 2:
            raise SystemExit(f"{chemin.name} : moins de deux VL exploitables.")
        reglage = reglages_pour(chemin)
        rendements = [couples[i][1] / couples[i - 1][1] - 1 for i in range(1, len(couples))]
        moyenne = sum(rendements) / len(rendements)
        variance = sum((x - moyenne) ** 2 for x in rendements) / len(rendements)
        vol_annuelle = (variance ** 0.5) * math.sqrt(252)
        pire_var = max(abs(x) for x in rendements)
        donnees = {
            "id": chemin.stem,
            "nom": reglage.get("nom", nom_du_fonds(chemin)),
            "type": reglage.get("type", "fonds"),
            "entree": reglage.get("entree", 0.0),
            "sortie": reglage.get("sortie", 0.0),
            "volAnnuelle": round(vol_annuelle, 4),
            "pireVarJour": round(pire_var, 4),
            "labels": [d.isoformat() for d, _ in couples],
            "vl": [v for _, v in couples],
        }
        if "politique" in reglage:
            donnees["politique"] = reglage["politique"]
        if reglage.get("rendement_du") and reglage.get("rendement_au"):
            b0, b1 = parser_date(reglage["rendement_du"]), parser_date(reglage["rendement_au"])
            i0 = max((i for i, (d, _) in enumerate(couples) if d <= b0), default=None)
            i1 = max((i for i, (d, _) in enumerate(couples) if d <= b1), default=None)
            if i0 is None or i1 is None or i1 <= i0:
                raise SystemExit(f"{chemin.name} : fenêtre de rendement invalide "
                                 f"({reglage['rendement_du']} → {reglage['rendement_au']}).")
            annees_f = (couples[i1][0] - couples[i0][0]).days / 365.25
            donnees["rendFige"] = {
                "r": round((couples[i1][1] / couples[i0][1]) ** (1 / annees_f) - 1, 6),
                "du": couples[i0][0].isoformat(),
                "au": couples[i1][0].isoformat(),
                "libelle": reglage.get("rendement_libelle", "la période retenue"),
            }
        fonds.append({
            "fichier": chemin.name,
            "ordre": reglage.get("ordre", 99),
            "donnees": donnees,
        })

    noms_observes = {f["donnees"]["nom"].casefold() for f in fonds}
    for c in FONDS_CIBLES:
        if c["nom"].casefold() in noms_observes:
            print(f"  (i) {c['nom']} : un historique de VL existe désormais — "
                  "entrée FONDS_CIBLES ignorée, pensez à la retirer.")
            continue
        ident = unicodedata.normalize("NFKD", c["nom"]).encode("ascii", "ignore").decode().lower()
        fonds.append({
            "fichier": None,
            "ordre": c.get("ordre", 50),
            "donnees": {
                "id": "cible-" + re.sub(r"[^a-z0-9]+", "-", ident).strip("-"),
                "nom": c["nom"],
                "type": "cible",
                "politique": c["politique"],
                "tauxCible": round(c["taux"] / 100.0, 6),
                "cibleAnnee": c["annee"],
                "entree": c.get("entree", 0.0),
                "sortie": c.get("sortie", 0.0),
                "volAnnuelle": None,
                "pireVarJour": None,
                "labels": [],
                "vl": [],
            },
        })

    fonds.sort(key=lambda f: (f["ordre"], f["donnees"]["nom"]))
    liste = [f["donnees"] for f in fonds]

    gabarit = GABARIT.read_text(encoding="utf-8")
    for repere in ("__FONDSDATA__", "__EMAILJSCONF__"):
        if repere not in gabarit:
            raise SystemExit(f"Repère {repere} introuvable dans le gabarit.")
    SORTIE.write_text(
        gabarit.replace("__FONDSDATA__", json.dumps(liste, ensure_ascii=False))
        .replace("__EMAILJSCONF__", json.dumps(EMAILJS)),
        encoding="utf-8",
    )

    print(f"{SORTIE.name} engendré — {len(liste)} fonds :")
    for f in liste:
        if f["type"] == "cible":
            print(
                f"  · {f['nom']} : rendement cible {f['tauxCible'] * 100:+.2f} % "
                f"(annoncé pour {f['cibleAnnee']}), sans historique de VL — "
                f"droits {f['entree']} % / {f['sortie']} %"
            )
            continue
        j0, j1 = date.fromisoformat(f["labels"][0]), date.fromisoformat(f["labels"][-1])
        annees = (j1 - j0).days / 365.25
        r = (f["vl"][-1] / f["vl"][0]) ** (1 / annees) - 1
        print(
            f"  · {f['nom']} : {len(f['vl'])} VL, du {j0} au {j1} "
            f"({(j1 - j0).days} j), rendement annualisé {r * 100:+.2f} %, "
            f"volatilité {f['volAnnuelle'] * 100:.1f} %, pire jour ±{f['pireVarJour'] * 100:.2f} %"
        )
        if "rendFige" in f:
            rf = f["rendFige"]
            print(
                f"      rendement appliqué : {rf['r'] * 100:+.2f} % — fenêtre "
                f"{rf['du']} → {rf['au']} ({rf['libelle']}) ; "
                f"historique complet {r * 100:+.2f} % pour mémoire"
            )


if __name__ == "__main__":
    main()
